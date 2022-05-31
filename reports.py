from openpyxl import load_workbook

from database_operation import get_bon_by_month, get_operations_by_commande_id
from tools import un_forming_date


def program_report(data):
    wb = load_workbook('xslx/program.xlsx')
    ws = wb.active

    menu = data[0]
    month = data[1]
    year = data[2]

    print(menu)
    print(month)
    print(year)

    for i in range(11, 18):
        day = menu[i-11]
        print(day)
        ws["F" + str(i)] = day[0]
        ws["E" + str(i)] = day[1]
        ws["D" + str(i)] = day[2]
        ws["C" + str(i)] = day[3]
        ws["B" + str(i)] = day[4]
        ws["A" + str(i)] = day[5]

    ws["D8"] = month + "/" + year

    wb.save("xslx/program.xlsx")


def sortie_report(data):
    wb = load_workbook('xslx/sortie_model.xlsx')
    ws = wb.active

    sortie = data[0]
    opertations = data[1]

    date = un_forming_date(sortie[2])

    ws["D8"] = date
    ws["E10"] = sortie[3]

    i = 0
    for operation in opertations:
        index = 13+i
        ws["G"+str(int(index))] = i + 1
        if operation[2] == "no_unit":
            prod = str(operation[1])
        else:
            prod = str(operation[1]) + operation[2]

        ws["D" + str(index)] = operation[0]
        ws["B" + str(index)] = prod

        i = i+1


    index = 13 + i

    ws.delete_rows(index, 47 - index)

    wb.save("xslx/raports/طلبية التموين.xlsx")


def entree_mois_report(data):
    wb = load_workbook('xslx/entree_mois_model.xlsx')
    ws = wb.active

    month = data[1] + " / " + str(data[2])

    ws["D7"] = month

    bons = get_bon_by_month("commande", data)

    prods = []
    for bon in bons:
        opertaions = get_operations_by_commande_id(bon[0])
        for operation in opertaions:
            go = True
            prod = []
            name = operation[0]
            for i in range(len(prods)):
                p = prods[i]
                if p[0] == name:
                    go = False
                    p[1] = p[1] + operation[1]
                    prods[i] = p
            if go:
                prod.append(operation[0])
                prod.append(operation[1])
                prod.append(operation[2])
                prods.append(prod)


    print(prods)


    for i in range(len(prods)):
        print("indeeeeeeeeeeeeeeeeeex", i)
        index = 11 + i
        prod = prods[i]
        ws["G" + str(index)] = i + 1
        ws["D" + str(index)] = prod[0]
        if prod[2] == "no_unit":
            ws["A" + str(index)] = str(prod[1])
        else:
            ws["A" + str(index)] = str(prod[1]) + prod[2]

        ind = 11 + i

    ind = ind + 1

    ws.delete_rows(ind, 99 - ind)


    wb.save("xslx/raports/تقرير المدخولات الشهري.xlsx")
    

    



def sortie_mois_report(data):
    wb = load_workbook('xslx/sortie_mois_model.xlsx')
    ws = wb.active


    wb.save("xslx/raports/تقرير التموين الشهري.xlsx")

def entree_year_report(year):
    wb = load_workbook('xslx/entree_year_model.xlsx')
    ws = wb.active


    wb.save("xslx/raports/تقرير المدخولات السنوي.xlsx")

def sortie_year_report(year):
    wb = load_workbook('xslx/sortie_year_model.xlsx')
    ws = wb.active


    wb.save("xslx/raports/تقرير التموين السنوي.xlsx")





